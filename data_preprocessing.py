# ECE 6607 PROJECT
# 2022.11.30
import csv
import openpyxl
import re

filename = "./data/data_1.xlsx"
post_net_csv = "./data/post_net.csv"
pre_net_csv = "./data/pre_net.csv"


def initialize(workbook):
    if workbook["Sheet1"] is None:
        exit()
    if workbook["post-networking"] is None:
        workbook.create_sheet("post-networking")
    if workbook["pre-networking"] is None:
        workbook.create_sheet("pre-networking")
    if workbook["post-cluster"] is None:
        workbook.create_sheet("post-cluster")
    if workbook["pre-cluster"] is None:
        workbook.create_sheet("pre-cluster")
    workbook["post-cluster"]["A1"] = "Academic"
    workbook["post-cluster"]["B1"] = "Food"
    workbook["post-cluster"]["C1"] = "Transportation"
    workbook["post-cluster"]["D1"] = "Infrastructure"
    workbook["post-cluster"]["E1"] = "Work"
    workbook["post-cluster"]["F1"] = "Finance"
    workbook["post-cluster"]["G1"] = "Health"
    workbook["post-cluster"]["H1"] = "Other"

    workbook["pre-cluster"]["A1"] = "Academic"
    workbook["pre-cluster"]["B1"] = "Food"
    workbook["pre-cluster"]["C1"] = "Transportation"
    workbook["pre-cluster"]["D1"] = "Infrastructure"
    workbook["pre-cluster"]["E1"] = "Work"
    workbook["pre-cluster"]["F1"] = "Finance"
    workbook["pre-cluster"]["G1"] = "Health"
    workbook["pre-cluster"]["H1"] = "Other"
    workbook.save(filename)


def splitData(workbook):
    #   Split raw data in xlsx file column into two different columns

    raw_data = workbook["Sheet1"]
    post_data = workbook["post-networking"]
    pre_data = workbook["pre-networking"]

    lines = list(raw_data.values)

    for l in range(1, 71):
        line = lines[l]
        foundFlag = False
        for i in range(len(line[1])):
            # If this guy have both pre- and post-networking data:
            if line[1][i] is '\n' and line[1][i + 1] is '\n':
                # print("line " + str(l) + " has both pre- and post-networking data!")
                post_part = line[1][0:i]
                pre_part = line[1][i+2:]
                post_data.cell(l, openpyxl.utils.column_index_from_string("A"), value=post_part)
                pre_data.cell(l, openpyxl.utils.column_index_from_string("A"), value=pre_part)
                foundFlag = True
                break
        #   If the guy only submitted post networking data:
        if foundFlag is False:
            # print("line " + str(l) + " only have post-networking data!")
            post_data.cell(l, openpyxl.utils.column_index_from_string("A"), value=line[1])

    workbook.save(filename)


def clusterData(sheet, clusters, cluster_sheet, counter):
    #   Cluster the suggestions into several predefined group based on keywords,
    #   and put all other suggestions don't match any keywords into a single group waiting for manually processing

    lines = list(sheet.values)
    for l in range(70):
        line = lines[l]
        if line[0] is not None:
            suggestions = line[0]
            newline = [0]
            for i in range(len(suggestions)):
                if suggestions[i] is '\n':
                    newline.append(i)
            # print("line " + str(l) + " " + str(newline))

            for j in range(len(newline)):
                if j == 0:
                    suggest = suggestions[newline[j]:newline[j + 1]]
                elif j == len(newline) - 1:
                    suggest = suggestions[newline[j] + 1:]
                else:
                    suggest = suggestions[newline[j] + 1:newline[j + 1]]
                #     One suggestion can match more than one clusters
                matchFlag = False
                for key in clusters.keys():
                    keywords = clusters[key]
                    for keyword in keywords:
                        if keyword == "RA" or keyword == "TA" or keyword == "GPA":
                            regrex = re.compile(keyword)
                        else:
                            regrex = re.compile(keyword, re.IGNORECASE)
                        if regrex.search(suggest) is not None:
                            cluster_sheet.cell(counter[key]+2, openpyxl.utils.column_index_from_string(key), value=suggest)
                            # print("suggest " + str(suggest) + " match to class " + str(key))
                            matchFlag = True
                            counter[key] += 1
                            break
                if matchFlag is False:
                    cluster_sheet.cell(counter["H"]+2, openpyxl.utils.column_index_from_string("H"), value=suggest)
                    counter["H"] += 1
                    # print("suggest " + str(suggest) + " match to class Other")


def xlsx2csv(sheet, csvfile):
    with open(csvfile, "w", newline='') as f:
        writer = csv.writer(f)
        data = []
        for i in range(1, sheet.max_row+1):
            row_stack = []
            for j in range(1, sheet.max_column+1):
                row_stack.append(sheet.cell(row=i, column=j).value)
            data.append(row_stack)
        writer.writerows(data)


def preprocessData():
    # create sheets, split data, cluster data, generate csv for visualization

    #   open xlsx file and initialize it
    xlsx_file = openpyxl.load_workbook(filename)
    initialize(xlsx_file)
    post_data = xlsx_file["post-networking"]
    pre_data = xlsx_file["pre-networking"]
    post_cluster_data = xlsx_file["post-cluster"]
    pre_cluster_data = xlsx_file["pre-cluster"]

    #   split data into pre- and post-networking groups
    splitData(xlsx_file)

    #   TODO: predefined clusters & hot keywords
    #  "Academic":               "academic", "course", "class", "grade", "GPA", "credit", "exam", "test", "homework"
    #  "Food":                   "food", "dining"
    #  "Transportation":         "transport", "parking", "bike", "bus"
    #  "Infrastructure":         "library", "gym", "amenities", "housing", "wifi", "network"
    #  "Work":                   "job", "RA", "TA", "work"
    #  "Finance"                 "money", "tuition", "fee"
    #  "Health"                  "mental", "dental", "health"
    #  "Other":                  Not in the keywords
    clusters = {"A": ["academic", "course", "class", "grade", "GPA", "credit", "exam", "test", "homework"],
                "B": ["food", "dining"],
                "C": ["transport", "parking", "bike", "bus"],
                "D": ["library", "gym", "amenities", "housing", "wifi", "network"],
                "E": ["job", "RA", "TA", "work"],
                "F": ["money", "tuition", "fee"],
                "G": ["mental", "dental", "health"]
                }
    post_counter = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0, "G": 0, "H": 0}
    pre_counter = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0, "G": 0, "H": 0}

    #   cluster data into several groups and count the members of each group
    clusterData(post_data, clusters, post_cluster_data, post_counter)
    clusterData(pre_data, clusters, pre_cluster_data, pre_counter)
    xlsx_file.save(filename)
    print("pre_counter " + str(pre_counter))
    print("post_counter " + str(post_counter))

    #   trans xlsx to csv format
    xlsx2csv(pre_cluster_data, pre_net_csv)
    xlsx2csv(post_cluster_data, post_net_csv)


if __name__ == '__main__':
    preprocessData()
